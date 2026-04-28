#!/usr/bin/env python3
"""
AI Opportunity Map — Excel to JSON extractor
=============================================
Reads the structured .xlsx workbook and produces the JSON format
consumed by the HTML visualisation tool.

Usage:
    python3 extract_from_excel.py <input.xlsx> [output.json]

If output path is omitted, writes to <input_stem>.json in the same directory.

Requirements:
    pip install openpyxl
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from core import assemble


class XlsxAccessor:
    """Wraps an openpyxl workbook to provide the cell(sheet, row, col) interface."""

    def __init__(self, path: Path):
        self._wb = load_workbook(str(path), data_only=True)

    def cell(self, sheet_name: str, row: int, col: int):
        ws = self._wb[sheet_name]
        return ws.cell(row=row, column=col).value


def main():
    if len(sys.argv) < 2:
        print(f"Usage: python3 {Path(__file__).name} <input.xlsx> [output.json]")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else xlsx_path.with_suffix(".json")

    print(f"Reading: {xlsx_path}", file=sys.stderr)
    try:
        accessor = XlsxAccessor(xlsx_path)
        data = assemble(accessor, source_label=xlsx_path.name)
    except Exception as e:
        print(f"Extraction failed: {e}", file=sys.stderr)
        sys.exit(1)

    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))

    fc = data["framework_config"]
    print(f"Done: {out_path}", file=sys.stderr)
    print(f"  Client:             {data['meta']['client']}", file=sys.stderr)
    print(f"  Impact factors:     {len(fc['impact_factors'])}", file=sys.stderr)
    print(f"  Readiness factors:  {len(fc['readiness_factors'])}", file=sys.stderr)
    print(f"  Enablers:           {len(data['implementation_enablers'])}", file=sys.stderr)
    print(f"  Use cases:          {len(data['use_cases'])}", file=sys.stderr)


if __name__ == "__main__":
    main()
