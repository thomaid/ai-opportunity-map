#!/usr/bin/env python3
"""
AI Opportunity Map — Excel to JSON extractor
=============================================
Reads the structured workbook and produces the JSON format
consumed by the HTML visualisation tool.

Usage:
    python3 extract_to_json.py <input.xlsx> [output.json]

If output path is omitted, writes to <input_stem>.json in the same directory.
"""

import sys
import json
import re
from pathlib import Path
from openpyxl import load_workbook


# ── Layout constants (positions that are fixed by design) ─────────────────────
# Config tab
CFG_ENGAGEMENT_NAME_CELL = (4, 4)   # row, col  (C = col 3, but label in col 3, value in col 4)
CFG_METADATA_IF_COUNT    = (13, 4)  # D13 — number of impact factors
CFG_METADATA_RF_COUNT    = (14, 4)  # D14 — number of readiness factors
CFG_METADATA_EN_COUNT    = (15, 4)  # D15 — number of enablers
CFG_METADATA_UC_COUNT    = (16, 4)  # D16 — number of use cases
CFG_IF_DATA_ROW_START    = 22       # First impact factor data row
CFG_RF_SECTION_SEARCH_FROM = 30     # Scan from here for RF table

# Enablers tab
EN_RF_ID_ROW    = 6   # Row containing RF IDs (machine-readable)
EN_DATA_ROW_START = 8
EN_COL_ID       = 2   # B
EN_COL_NAME     = 3   # C
EN_COL_CATEGORY = 4   # D
EN_COL_EFFORT   = 5   # E
EN_COL_LEVEL    = 6   # F
EN_RF_COL_START = 9   # I  (first dynamic RF uplift column)

# Use Cases tab
UC_DATA_ROW_START = 7
UC_COL_ID         = 2   # B
UC_COL_NAME       = 3   # C
UC_COL_DOMAIN     = 4   # D
UC_COL_THEME      = 5   # E
UC_COL_STATE      = 6   # F
UC_COL_RISK       = 7   # G
# col 8 = Notes (not exported)
UC_COL_ENABLES    = 9   # I
UC_COL_ENABLERS   = 10  # J

# Score tabs
SC_FACTOR_ID_ROW  = 5   # Row containing factor IDs
SC_WEIGHT_ROW     = 7   # Row containing weights
SC_DATA_ROW_START = 8   # First use-case data row
SC_COL_UC_ID      = 2   # B
SC_COL_SCORE_START = 5  # E  (first dynamic factor column)


# ── Helpers ───────────────────────────────────────────────────────────────────

def cell(ws, row, col):
    """Return cell value, None if empty."""
    v = ws.cell(row=row, column=col).value
    return v


def parse_pipe(value):
    """Split a pipe-delimited string into a list, ignoring blanks."""
    if not value:
        return []
    return [s.strip() for s in str(value).split("|") if s.strip()]


def find_rf_data_start(ws_cfg, from_row, n_rf):
    """
    Scan down from from_row to find the first row of the RF data table.
    We identify it as the first row below from_row where col C matches RF[0-9]+.
    """
    for row in range(from_row, from_row + 40):
        v = cell(ws_cfg, row, 3)
        if v and re.match(r"RF\d+", str(v)):
            return row
    raise ValueError(f"Could not find Readiness Factor data starting from row {from_row}")


def read_factors(ws_cfg, data_row_start, n):
    """Read n factor rows from Config. Returns list of dicts."""
    factors = []
    for i in range(n):
        row = data_row_start + i
        fid   = cell(ws_cfg, row, 3)
        fname = cell(ws_cfg, row, 4)
        wt    = cell(ws_cfg, row, 5)
        desc  = cell(ws_cfg, row, 6)
        if not fid or not str(fid).strip():
            break
        factors.append({
            "id":          str(fid).strip(),
            "name":        str(fname).strip() if fname else "",
            "weight":      float(wt) if wt is not None else 1.0,
            "description": str(desc).strip() if desc else "",
        })
    return factors


def read_score_tab(ws, n_factors, uc_ids):
    """
    Read a score tab (Impact Scores or Readiness Scores).
    Returns:
      factor_ids  — ordered list of factor IDs from row 5
      weights     — dict {factor_id: weight} from row 7
      scores      — dict {uc_id: {factor_id: score}}
    """
    factor_ids = []
    for col in range(SC_COL_SCORE_START, SC_COL_SCORE_START + n_factors + 10):
        v = cell(ws, SC_FACTOR_ID_ROW, col)
        if v and re.match(r"[IR]F\d+", str(v)):
            factor_ids.append(str(v).strip())
        elif factor_ids:
            break

    if len(factor_ids) < n_factors:
        raise ValueError(
            f"Expected {n_factors} factor IDs in row {SC_FACTOR_ID_ROW}, "
            f"found {len(factor_ids)}: {factor_ids}"
        )
    factor_ids = factor_ids[:n_factors]

    weights = {}
    for i, fid in enumerate(factor_ids):
        col = SC_COL_SCORE_START + i
        w = cell(ws, SC_WEIGHT_ROW, col)
        weights[fid] = float(w) if w is not None else 1.0

    scores = {}
    uc_id_set = set(uc_ids)
    for row in range(SC_DATA_ROW_START, SC_DATA_ROW_START + len(uc_ids) + 50):
        uc_id = cell(ws, row, SC_COL_UC_ID)
        if not uc_id:
            continue
        uc_id = str(uc_id).strip()
        if uc_id not in uc_id_set:
            continue
        row_scores = {}
        for i, fid in enumerate(factor_ids):
            col = SC_COL_SCORE_START + i
            v = cell(ws, row, col)
            row_scores[fid] = int(v) if v is not None else 3
        scores[uc_id] = row_scores

    return factor_ids, weights, scores


# ── Main extractor ────────────────────────────────────────────────────────────

def extract(xlsx_path: Path) -> dict:
    wb = load_workbook(str(xlsx_path), data_only=True)

    ws_cfg = wb["Config"]
    ws_en  = wb["Enablers"]
    ws_uc  = wb["Use Cases"]
    ws_is  = wb["Impact Scores"]
    ws_rs  = wb["Readiness Scores"]

    engagement_name = cell(ws_cfg, CFG_ENGAGEMENT_NAME_CELL[0], CFG_ENGAGEMENT_NAME_CELL[1]) or ""

    n_if = int(cell(ws_cfg, CFG_METADATA_IF_COUNT[0],  CFG_METADATA_IF_COUNT[1])  or 0)
    n_rf = int(cell(ws_cfg, CFG_METADATA_RF_COUNT[0],  CFG_METADATA_RF_COUNT[1])  or 0)
    n_en = int(cell(ws_cfg, CFG_METADATA_EN_COUNT[0],  CFG_METADATA_EN_COUNT[1])  or 0)
    n_uc = int(cell(ws_cfg, CFG_METADATA_UC_COUNT[0],  CFG_METADATA_UC_COUNT[1])  or 0)

    if not all([n_if, n_rf, n_en, n_uc]):
        raise ValueError(
            f"Metadata counts incomplete: IF={n_if}, RF={n_rf}, EN={n_en}, UC={n_uc}. "
            "Check the Metadata section in the Config tab."
        )

    impact_factors = read_factors(ws_cfg, CFG_IF_DATA_ROW_START, n_if)
    if len(impact_factors) != n_if:
        print(f"  NOTE: Metadata says {n_if} impact factors; found {len(impact_factors)}. "
              f"Using actual count.", file=sys.stderr)
        n_if = len(impact_factors)

    rf_data_start = find_rf_data_start(ws_cfg, CFG_RF_SECTION_SEARCH_FROM, n_rf)
    readiness_factors = read_factors(ws_cfg, rf_data_start, n_rf)
    if len(readiness_factors) != n_rf:
        print(f"  NOTE: Metadata says {n_rf} readiness factors; found {len(readiness_factors)}. "
              f"Using actual count.", file=sys.stderr)
        n_rf = len(readiness_factors)

    if_ids = [f["id"] for f in impact_factors]
    rf_ids = [f["id"] for f in readiness_factors]

    rf_col_map = {}
    for col in range(EN_RF_COL_START, EN_RF_COL_START + n_rf + 10):
        v = cell(ws_en, EN_RF_ID_ROW, col)
        if v and re.match(r"RF\d+", str(v)):
            rf_col_map[str(v).strip()] = col
        elif rf_col_map:
            break

    missing_rf = set(rf_ids) - set(rf_col_map.keys())
    if missing_rf:
        raise ValueError(f"Enablers tab missing RF columns for: {missing_rf}")

    enablers = []
    for i in range(200):
        row = EN_DATA_ROW_START + i
        en_id = cell(ws_en, row, EN_COL_ID)
        if not en_id or not str(en_id).strip():
            break
        en_id = str(en_id).strip()
        if not re.match(r"EN\d+", en_id):
            break

        factor_uplifts = []
        for rf_id in rf_ids:
            col = rf_col_map.get(rf_id)
            if col is None:
                continue
            v = cell(ws_en, row, col)
            if v is not None and str(v).strip() not in ("", "0"):
                try:
                    uplift = int(v)
                    if uplift > 0:
                        factor_uplifts.append({"factor_id": rf_id, "max_uplift": uplift})
                except (ValueError, TypeError):
                    pass

        enablers.append({
            "id":               en_id,
            "name":             str(cell(ws_en, row, EN_COL_NAME) or "").strip(),
            "category":         str(cell(ws_en, row, EN_COL_CATEGORY) or "").strip(),
            "effort_scale":     str(cell(ws_en, row, EN_COL_EFFORT) or "").strip(),
            "completion_level": int(cell(ws_en, row, EN_COL_LEVEL) or 0),
            "factor_uplifts":   factor_uplifts,
        })

    if len(enablers) != n_en:
        print(f"  NOTE: Metadata says {n_en} enablers; found {len(enablers)} data rows. "
              f"Metadata may be stale — using actual row count.", file=sys.stderr)

    use_cases_raw = []
    for i in range(500):
        row = UC_DATA_ROW_START + i
        uc_id = cell(ws_uc, row, UC_COL_ID)
        if not uc_id or not str(uc_id).strip():
            break
        if not re.match(r"UC\d+", str(uc_id).strip()):
            break
        use_cases_raw.append({
            "id":                          str(uc_id).strip(),
            "name":                        str(cell(ws_uc, row, UC_COL_NAME)   or "").strip(),
            "business_domain":             str(cell(ws_uc, row, UC_COL_DOMAIN) or "").strip(),
            "investment_theme":            str(cell(ws_uc, row, UC_COL_THEME)  or "").strip(),
            "current_state":               str(cell(ws_uc, row, UC_COL_STATE)  or "").strip(),
            "regulatory_risk_classification": str(cell(ws_uc, row, UC_COL_RISK) or "").strip(),
            "enables":                     parse_pipe(cell(ws_uc, row, UC_COL_ENABLES)),
            "implementation_enablers":     parse_pipe(cell(ws_uc, row, UC_COL_ENABLERS)),
        })

    if len(use_cases_raw) != n_uc:
        print(f"  NOTE: Metadata says {n_uc} use cases; found {len(use_cases_raw)}. "
              f"Using actual count.", file=sys.stderr)
    uc_ids = [u["id"] for u in use_cases_raw]
    uc_id_set = set(uc_ids)

    _, if_weights, impact_scores   = read_score_tab(ws_is, n_if, uc_ids)
    _, rf_weights, readiness_scores = read_score_tab(ws_rs, n_rf, uc_ids)

    for f in impact_factors:
        if f["id"] in if_weights:
            f["weight"] = if_weights[f["id"]]
    for f in readiness_factors:
        if f["id"] in rf_weights:
            f["weight"] = rf_weights[f["id"]]

    use_cases = []
    for uc in use_cases_raw:
        uid = uc["id"]
        i_scores = impact_scores.get(uid, {fid: 3 for fid in if_ids})
        r_scores = readiness_scores.get(uid, {rfid: 3 for rfid in rf_ids})

        bad_enables  = [x for x in uc["enables"] if x not in uc_id_set]
        bad_enablers = [x for x in uc["implementation_enablers"] if x not in {e["id"] for e in enablers}]
        if bad_enables:
            print(f"  WARNING {uid}: enables references unknown UC IDs: {bad_enables}", file=sys.stderr)
        if bad_enablers:
            print(f"  WARNING {uid}: implementation_enablers references unknown EN IDs: {bad_enablers}", file=sys.stderr)

        use_cases.append({
            **uc,
            "impact_scores":    i_scores,
            "readiness_scores": r_scores,
        })

    return {
        "meta": {
            "version":          "1.0",
            "source":           xlsx_path.name,
            "client":           str(engagement_name).strip(),
            "data_model_version": "0.2",
        },
        "framework_config": {
            "impact_factors":    impact_factors,
            "readiness_factors": readiness_factors,
        },
        "implementation_enablers": enablers,
        "use_cases":               use_cases,
    }


# ── CLI entry point ───────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(f"Usage: python3 {Path(__file__).name} <input.xlsx> [output.json]")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else xlsx_path.with_suffix(".json")

    print(f"Reading: {xlsx_path}", file=sys.stderr)
    try:
        data = extract(xlsx_path)
    except Exception as e:
        print(f"Extraction failed: {e}", file=sys.stderr)
        sys.exit(1)

    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))

    fc = data["framework_config"]
    print(f"Done: {out_path}", file=sys.stderr)
    print(f"  Client:             {data['meta']['client']}", file=sys.stderr)
    print(f"  Impact factors:     {len(fc['impact_factors'])}", file=sys.stderr)
    print(f"  Readiness factors:  {len(fc['readiness_factors'])}", file=sys.stderr)
    print(f"  Enablers:           {len(data['implementation_enablers'])}", file=sys.stderr)
    print(f"  Use cases:          {len(data['use_cases'])}", file=sys.stderr)


if __name__ == "__main__":
    main()
